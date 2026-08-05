import streamlit as st
import os, io, json, re
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    st.error("pip install openpyxl"); st.stop()

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import (BaseDocTemplate, Frame, PageTemplate,
                                     Table, TableStyle, Paragraph, Spacer,
                                     Image as RLImage, HRFlowable, KeepInFrame)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# Logo dosyadan oku
LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.jpeg")
with open(LOGO_PATH, "rb") as f:
    LOGO_BYTES = f.read()

FIRMA_BILGI_TXT = (
    "TA Test Analiz Sistemleri ve Kimyasal Maddeler Sanayi ve Ticaret Limited Şirketi\n"
    "Monumento Plaza d:120 Esentepe 34870 Kartal İSTANBUL tel: 0216 546 1095\n"
    "Muammer Aksoy Cad. No:81 d:126 Altınşehir 16230 Nilüfer BURSA tel: 0224 502 6203\n"
    "www.takimya.com"
)
FIRMA_HTML = (
    "<b>TA Test Analiz Sistemleri ve Kimyasal Maddeler Sanayi ve Ticaret Limited Şirketi</b><br/>"
    "Monumento Plaza d:120 Esentepe 34870 Kartal İSTANBUL tel: 0216 546 1095<br/>"
    "Muammer Aksoy Cad. No:81 d:126 Altınşehir 16230 Nilüfer BURSA tel: 0224 502 6203<br/>"
    "www.takimya.com"
)
EURO_FMT = "#,##0.00\\ [$\u20ac-1]"

AYARLAR_DOSYA = os.path.join(os.path.dirname(__file__), "ayarlar.json")

def ayarlari_yukle():
    v = {"iletisim_ad": "", "iletisim_tel": "", "iletisim_email": ""}
    try:
        if os.path.exists(AYARLAR_DOSYA):
            with open(AYARLAR_DOSYA, encoding="utf-8") as f:
                v.update(json.load(f))
    except Exception:
        pass
    return v

def ayarlari_kaydet(a):
    try:
        with open(AYARLAR_DOSYA, "w", encoding="utf-8") as f:
            json.dump(a, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def urunu_ayir(metin: str) -> list:
    metin = metin.replace("\r\n", "\n").replace("\r", "\n")
    for marker in ["---- Original Message ----", "----İletilen İleti----", "----Forwarded"]:
        idx = metin.find(marker)
        if idx != -1:
            bolum = metin[idx:]
            subj_m = re.search(r"Subject:.*\n", bolum)
            metin = bolum[subj_m.end():] if subj_m else bolum
            break

    bloklar = re.compile(r"={6,}").split(metin)

    # Fiyat satırından önceki boş satırları her iki formatta da kaldır
    metin2 = re.sub(r"\n{2,}([\d]{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?|[\d]+(?:[.,]\d{2})?)\s*(?:EUR|€)",
                    r"\n\1 EUR", metin)
    bloklar = re.compile(r"={6,}").split(metin2)

    # ======= yoksa boş satırla böl
    if len(bloklar) <= 2:
        bloklar = re.split(r"\n{2,}", metin2)

    urunler = []
    for blok in bloklar:
        blok = blok.strip()
        if not blok: continue
        satirlar = [s.strip() for s in blok.splitlines() if s.strip()]
        if not satirlar: continue
        fiyat = None; fiyat_idx = None
        for i in range(len(satirlar) - 1, -1, -1):
            # Görünmez karakterleri temizle
            satir = satirlar[i].strip().replace('\xa0', ' ').replace('\u202f', ' ')
            m = re.fullmatch(r"\s*([\d]{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?|[\d]+(?:[.,]\d{2})?)\s*(?:EUR|€)\s*", satir)
            if not m:
                m = re.match(r"^([\d]{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?|[\d]+(?:[.,]\d{2})?)\s*(?:EUR|€)\s*$", satir)
            if m:
                fiyat_idx = i; raw = m.group(1)
                if "," in raw and "." in raw:
                    raw = (raw.replace(".", "").replace(",", ".")
                           if raw.index(".") < raw.index(",") else raw.replace(",", ""))
                elif "," in raw:
                    raw = raw.replace(",", ".")
                try: fiyat = float(raw)
                except: fiyat = 0.0
                break
        if fiyat is None: continue
        ic = satirlar[:fiyat_idx]
        if not ic: continue
        if any(skip in ic[0] for skip in
               ["Original Message", "İletilen", "Forwarded", "Fatih", "Sener", "takimya", "@"]):
            continue
        urunler.append({"kod": ic[0], "aciklama": "\n".join(ic), "fiyat": fiyat, "adet": 1})
    return urunler

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def excel_olustur(urunler, musteri_adi, musteri_sehir, teklif_no, teklif_tarihi,
                  teslim_suresi, teslimat, alt_notlar, iletisim, indirim, cikti):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    for col, w in [("A",7.11),("B",69.55),("C",5.66),("D",15.44),("E",15.66),("F",7.11)]:
        ws.column_dimensions[col].width = w
    ws.sheet_view.view = "normal"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left=0.4; ws.page_margins.right=0.48
    ws.page_margins.top=0.46; ws.page_margins.bottom=0.48
    ws.print_area = "A1:F100"

    b12  = Font(name="Calibri", size=12, bold=True)
    n12  = Font(name="Calibri", size=12)
    n11  = Font(name="Calibri", size=11)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ltop = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    ctop = Alignment(horizontal="center", vertical="top",    wrap_text=True)
    lmid = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Firma başlığı
    ws.merge_cells("B2:D9")
    ws["B2"].value = FIRMA_BILGI_TXT; ws["B2"].font = b12; ws["B2"].alignment = ctr
    ws.row_dimensions[2].height = 65

    # Logo
    try:
        from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
        from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
        from openpyxl.utils.units import pixels_to_EMU
        img = XLImage(io.BytesIO(LOGO_BYTES))
        img.width = 220; img.height = 56
        pos  = XDRPoint2D(pixels_to_EMU(195), pixels_to_EMU(6))
        size = XDRPositiveSize2D(pixels_to_EMU(220), pixels_to_EMU(56))
        img.anchor = AbsoluteAnchor(pos=pos, ext=size)
        ws.add_image(img)
    except Exception:
        try:
            img2 = XLImage(io.BytesIO(LOGO_BYTES)); img2.width=220; img2.height=56; img2.anchor="C1"; ws.add_image(img2)
        except: pass

    # FİYAT TEKLİFİ
    ws.merge_cells("B12:D12")
    ws["B12"].value = "FİYAT TEKLİFİ"; ws["B12"].font = b12; ws["B12"].alignment = ctr
    ws.row_dimensions[12].height = 20

    # Teklif No / Tarih / Teslim Süresi
    ws.merge_cells("C15:D15")
    ws["C15"].value = "Teklif No :"; ws["C15"].font = b12; ws["C15"].alignment = lmid
    ws["E15"].value = teklif_no; ws["E15"].font = n12; ws["E15"].alignment = lmid
    ws.merge_cells("C16:D16")
    ws["C16"].value = "Tarih :"; ws["C16"].font = b12; ws["C16"].alignment = lmid
    ws["E16"].value = teklif_tarihi; ws["E16"].font = n12; ws["E16"].alignment = lmid
    if teslim_suresi.strip():
        ws.merge_cells("C17:D17")
        ws["C17"].value = "Teslim Süresi :"; ws["C17"].font = b12; ws["C17"].alignment = lmid
        ws["E17"].value = teslim_suresi; ws["E17"].font = n12; ws["E17"].alignment = lmid
        ws.row_dimensions[17].height = 16

    # Müşteri
    ws.merge_cells("A17:B17")
    ws["A17"].value = musteri_adi.upper(); ws["A17"].font = b12; ws["A17"].alignment = ltop
    ws.merge_cells("A18:B18")
    ws["A18"].value = musteri_sehir.upper(); ws["A18"].font = b12; ws["A18"].alignment = ltop

    # Tablo başlığı
    ws.row_dimensions[21].height = 30
    for col, val in [("A","No"),("B","AÇIKLAMA"),("C","ADET"),("D","BİRİM FİYAT"),("E","TOPLAM FİYAT")]:
        c = ws[f"{col}21"]; c.value = val; c.font = n12; c.alignment = ctr; c.border = _thin()

    def rb():
        from openpyxl.styles import Border as B, Side as S
        t = S(style="thin"); return B(left=t, right=t, top=t, bottom=t)

    DS = 22; r = DS
    for idx, urun in enumerate(urunler, 1):
        adet = urun.get("adet", 1)
        ws.row_dimensions[r].height = max(15, min(urun["aciklama"].count("\n")*13.8+13.8, 600))
        ws[f"A{r}"].value = idx;              ws[f"A{r}"].font=n12; ws[f"A{r}"].alignment=ctop; ws[f"A{r}"].border=rb()
        ws[f"B{r}"].value = urun["aciklama"]; ws[f"B{r}"].font=n11; ws[f"B{r}"].alignment=ltop; ws[f"B{r}"].border=rb()
        ws[f"C{r}"].value = adet;             ws[f"C{r}"].font=n12; ws[f"C{r}"].alignment=ctop
        ws[f"C{r}"].number_format = "#,##0";  ws[f"C{r}"].border=rb()
        # Birim fiyat D sütununda
        fv = urun["fiyat"]
        ws[f"D{r}"].value = int(fv) if fv==int(fv) else fv
        ws[f"D{r}"].font=n12; ws[f"D{r}"].alignment=ltop; ws[f"D{r}"].number_format=EURO_FMT; ws[f"D{r}"].border=rb()
        # Toplam fiyat = birim * adet
        ws[f"E{r}"].value = f"=D{r}*C{r}"
        ws[f"E{r}"].font=n12; ws[f"E{r}"].alignment=ltop; ws[f"E{r}"].number_format=EURO_FMT; ws[f"E{r}"].border=rb()
        r += 1

    # TOPLAM → İNDİRİM → NET TOPLAM
    tr = r + 1
    rmid = Alignment(horizontal="right", vertical="center")

    # 1) TOPLAM
    toplam_r = tr
    ws[f"D{tr}"].value = "TOPLAM"; ws[f"D{tr}"].font = b12; ws[f"D{tr}"].alignment = rmid
    ws[f"E{tr}"].value = f"=SUM(E{DS}:E{r-1})"
    ws[f"E{tr}"].font = b12; ws[f"E{tr}"].number_format = EURO_FMT
    tr += 1

    if indirim and indirim > 0:
        # 2) İNDİRİM
        indirim_r = tr
        ws[f"D{tr}"].value = "İNDİRİM"; ws[f"D{tr}"].font = b12; ws[f"D{tr}"].alignment = rmid
        ws[f"E{tr}"].value = indirim; ws[f"E{tr}"].font = b12; ws[f"E{tr}"].number_format = EURO_FMT
        tr += 1

        # 3) Net tutar
        net_r = tr
        ws[f"E{tr}"].value = f"=E{toplam_r}-E{indirim_r}"
        ws[f"E{tr}"].font = b12; ws[f"E{tr}"].number_format = EURO_FMT
        from openpyxl.styles import Border, Side as Sd
        top_border = Border(top=Sd(style="thin"))
        ws[f"E{tr}"].border = top_border
        tr += 1

    # Alt notlar + Teslimat (sol) | İletişim (sağ — D:E sütunu)
    nr = tr + 3
    # Sol: alt notlar
    sol_nr = nr
    for txt in alt_notlar:
        if txt.strip():
            ws[f"B{sol_nr}"].value = txt; ws[f"B{sol_nr}"].font = b12
            ws[f"B{sol_nr}"].alignment = Alignment(vertical="top", wrap_text=True)
            sol_nr += 2
    if teslimat.strip():
        ws[f"B{sol_nr}"].value = f"Teslimat: {teslimat}"; ws[f"B{sol_nr}"].font = b12
        ws[f"B{sol_nr}"].alignment = Alignment(vertical="top", wrap_text=True); sol_nr += 2

    # Sağ: iletişim — D:E sütunlarında, nr satırından başla
    if iletisim.strip():
        il_nr = nr
        ws[f"D{il_nr}"].value = "İletişim:"; ws[f"D{il_nr}"].font = b12
        ws[f"D{il_nr}"].alignment = Alignment(vertical="top", wrap_text=True); il_nr += 1
        for satir in iletisim.strip().splitlines():
            ws.merge_cells(f"D{il_nr}:E{il_nr}")
            ws[f"D{il_nr}"].value = satir; ws[f"D{il_nr}"].font = n12
            ws[f"D{il_nr}"].alignment = Alignment(vertical="top", wrap_text=True); il_nr += 1

    wb.save(cikti)
    return len(urunler)

def _reg_pdf_fonts():
    import glob
    # Mevcut fontları logla
    found = glob.glob("/usr/share/fonts/**/*.ttf", recursive=True)
    print("=== MEVCUT TTF FONTLAR ===")
    for f in sorted(found)[:30]:
        print(f)
    
    candidates = [
        # DejaVu — Streamlit Cloud'da mevcut, Türkçe destekler
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        # Liberation — Linux
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
        # Ubuntu font
        ("/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf",
         "/usr/share/fonts/truetype/ubuntu/Ubuntu-B.ttf"),
        # Windows
        (r"C:\Windows\Fonts\arial.ttf",   r"C:\Windows\Fonts\arialbd.ttf"),
        (r"C:\Windows\Fonts\calibri.ttf", r"C:\Windows\Fonts\calibrib.ttf"),
    ]
    for reg, bold in candidates:
        print(f"Deneniyor: {reg} -> exists={os.path.exists(reg)}")
        if os.path.exists(reg) and os.path.exists(bold):
            try:
                pdfmetrics.registerFont(TTFont("Sans",     reg))
                pdfmetrics.registerFont(TTFont("SansBold", bold))
                print(f"KULLANILAN FONT: {reg}")
                return ("Sans", "SansBold")
            except Exception as e:
                print(f"Font hatası: {e}")
                continue
    print("Helvetica kullanılıyor!")
    return ("Helvetica", "Helvetica-Bold")

def _fmt_euro(val):
    if val == 0: return "0,00 \u20ac"
    s = f"{val:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    return f"{s} \u20ac"

def pdf_olustur(urunler, musteri_adi, musteri_sehir, teklif_no, teklif_tarihi,
                teslim_suresi, teslimat, alt_notlar, iletisim, indirim, cikti):
    if not HAS_PDF:
        raise ImportError("pip install reportlab")
    FONT, FONT_BOLD = _reg_pdf_fonts()
    PW, PH = A4; MAR = 1.5*cm; CW = PW - 2*MAR
    doc = BaseDocTemplate(cikti, pagesize=A4,
                          leftMargin=MAR, rightMargin=MAR,
                          topMargin=MAR, bottomMargin=MAR)
    frame = Frame(MAR, MAR, CW, PH-2*MAR, id="normal")
    doc.addPageTemplates([PageTemplate(id="main", frames=frame)])
    story = []

    def sty(name, **kw):
        d = dict(fontName=FONT, fontSize=9, leading=12, spaceAfter=0, spaceBefore=0)
        d.update(kw); return ParagraphStyle(name, **d)
    def para(text, style):
        text = str(text).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace("\n","<br/>")
        return Paragraph(text, style)
    def hpara(html, style):
        return Paragraph(html, style)

    S_FIRMA = sty("firma", fontName=FONT_BOLD, fontSize=8, alignment=TA_CENTER, leading=13)
    S_BOLD  = sty("bold",  fontName=FONT_BOLD)
    S_ACK   = sty("ack",   fontSize=8, leading=11)
    S_SM    = sty("sm",    fontSize=8, leading=11)

    # Logo + başlık
    logo_img = RLImage(io.BytesIO(LOGO_BYTES), width=7*cm, height=1.8*cm)
    logo_img.hAlign = "CENTER"
    firma_p = Paragraph(FIRMA_HTML, S_FIRMA)
    hdr = Table([[logo_img], [firma_p]], colWidths=[CW])
    hdr.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("BOTTOMPADDING",(0,0),(0,0),4),
        ("BOTTOMPADDING",(0,1),(0,1),8),
    ]))
    story += [hdr,
              HRFlowable(width=CW, thickness=1, color=colors.black),
              Spacer(1, 3*mm),
              hpara("<b>FİYAT TEKLİFİ</b>",
                    sty("fth", fontName=FONT_BOLD, fontSize=11, alignment=TA_CENTER)),
              Spacer(1, 3*mm)]

    # Müşteri + Teklif No/Tarih/Teslim
    meta = [
        [hpara(f"<b>{musteri_adi.upper()}</b>", S_BOLD),
         hpara(f"<b>Teklif No :</b>  {teklif_no}", sty("mn1", fontName=FONT_BOLD, alignment=TA_RIGHT))],
        [hpara(f"<b>{musteri_sehir.upper()}</b>", S_BOLD),
         hpara(f"<b>Tarih :</b>  {teklif_tarihi}", sty("mn2", fontName=FONT_BOLD, alignment=TA_RIGHT))],
    ]
    if teslim_suresi.strip():
        meta.append(["", hpara(f"<b>Teslim Süresi :</b>  {teslim_suresi}", sty("mn3", fontName=FONT_BOLD, alignment=TA_RIGHT))])
    mt = Table(meta, colWidths=[CW*0.6, CW*0.4])
    mt.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("BOTTOMPADDING",(0,0),(-1,-1),1),("TOPPADDING",(0,0),(-1,-1),1)]))
    story += [mt, Spacer(1, 4*mm)]

    # Ürün tablosu
    C0=0.8*cm; C2=1.3*cm; C3=2.8*cm; C4=2.8*cm; C1=CW-C0-C2-C3-C4
    tbl = [[
        hpara("<b>No</b>",           sty("h0",fontName=FONT_BOLD,fontSize=8,alignment=TA_CENTER,leading=10)),
        hpara("<b>AÇIKLAMA</b>",     sty("h1",fontName=FONT_BOLD,fontSize=9,alignment=TA_CENTER,leading=11)),
        hpara("<b>ADET</b>",         sty("h2",fontName=FONT_BOLD,fontSize=8,alignment=TA_CENTER,leading=10)),
        hpara("<b>BİRİM FİYAT</b>",  sty("h3",fontName=FONT_BOLD,fontSize=8,alignment=TA_CENTER,leading=10)),
        hpara("<b>TOPLAM FİYAT</b>", sty("h4",fontName=FONT_BOLD,fontSize=8,alignment=TA_CENTER,leading=10)),
    ]]
    toplam = 0
    for i, urun in enumerate(urunler, 1):
        fv = urun["fiyat"]; adet = urun.get("adet",1); toplam += fv * adet
        tbl.append([
            para(str(i),             sty(f"n{i}", alignment=TA_CENTER, fontSize=8)),
            para(urun["aciklama"],   S_ACK),
            para(str(adet),          sty(f"a{i}", alignment=TA_CENTER, fontSize=8)),
            para(_fmt_euro(fv),      sty(f"b{i}", alignment=TA_RIGHT,  fontSize=8)),
            para(_fmt_euro(fv*adet), sty(f"p{i}", alignment=TA_RIGHT,  fontSize=8)),
        ])
    if indirim and indirim > 0:
        tbl.append(["","","",
            hpara("<b>TOPLAM</b>", sty("tt",fontName=FONT_BOLD,alignment=TA_CENTER)),
            hpara(f"<b>{_fmt_euro(toplam)}</b>", sty("tp",fontName=FONT_BOLD,alignment=TA_CENTER)),
        ])
        tbl.append(["","","",
            hpara("<b>İNDİRİM</b>", sty("ind2",fontName=FONT_BOLD,alignment=TA_CENTER)),
            hpara(f"<b>{_fmt_euro(indirim)}</b>", sty("indp2",fontName=FONT_BOLD,alignment=TA_CENTER)),
        ])
        toplam_net = toplam - indirim
        tbl.append(["","","","",
            hpara(f"<b>{_fmt_euro(toplam_net)}</b>", sty("net",fontName=FONT_BOLD,alignment=TA_CENTER)),
        ])
    else:
        toplam_net = toplam
        tbl.append(["","","",
            hpara("<b>TOPLAM</b>", sty("tt",fontName=FONT_BOLD,alignment=TA_CENTER)),
            hpara(f"<b>{_fmt_euro(toplam_net)}</b>", sty("tp",fontName=FONT_BOLD,alignment=TA_CENTER)),
        ])
    NR = len(tbl)
    t = Table(tbl, colWidths=[C0,C1,C2,C3,C4], repeatRows=1,
              rowHeights=[20]+[None]*(NR-1))
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#d9d9d9")),
        ("VALIGN",(0,0),(-1,0),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,0),5),("BOTTOMPADDING",(0,0),(-1,0),5),
        ("VALIGN",(0,1),(-1,-2),"TOP"),
        ("TOPPADDING",(0,1),(-1,-2),4),("BOTTOMPADDING",(0,1),(-1,-2),4),
        ("LEFTPADDING",(1,1),(1,-1),4),
        ("TOPPADDING",(0,-1),(-1,-1),5),("BOTTOMPADDING",(0,-1),(-1,-1),5),
        ("LINEABOVE",(0,-1),(-1,-1),1,colors.black),
        ("BOX",(0,0),(-1,-1),0.5,colors.black),
        ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
        *[("BACKGROUND",(0,r),(-1,r),colors.HexColor("#f7f7f7")) for r in range(2,NR-1,2)],
    ]))
    story += [t, Spacer(1, 5*mm)]

    # Alt notlar + Teslimat (sol) | İletişim (sağ) — yan yana tablo
    sol_items = []
    for txt in alt_notlar:
        if txt.strip():
            sol_items.append(hpara(f"<b>{txt}</b>", S_BOLD))
    if teslimat.strip():
        sol_items.append(Spacer(1, 2*mm))
        sol_items.append(hpara(f"<b>Teslimat:</b>  {teslimat}", S_BOLD))

    sag_items = []
    if iletisim.strip():
        sag_items.append(hpara("<b>İletişim:</b>", S_BOLD))
        for satir in iletisim.strip().splitlines():
            sag_items.append(para(satir, S_SM))

    # Eşit uzunlukta yap
    max_len = max(len(sol_items), len(sag_items), 1)
    while len(sol_items) < max_len: sol_items.append(Spacer(1, 1*mm))
    while len(sag_items) < max_len: sag_items.append(Spacer(1, 1*mm))

    from reportlab.platypus import KeepInFrame
    sol_frame  = KeepInFrame(CW*0.55, 200, sol_items)
    sag_frame  = KeepInFrame(CW*0.4,  200, sag_items)

    footer_tbl = Table([[sol_frame, sag_frame]], colWidths=[CW*0.62, CW*0.38])
    footer_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN",  (1,0), (1,0),  "LEFT"),
    ]))
    story.append(footer_tbl)

    doc.build(story)
    return len(urunler)


# ── UI ────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Fiyat Teklifi Oluşturucu | TA Test Analiz",
                   page_icon="📊", layout="wide")

col_logo, col_title = st.columns([1, 4])
with col_logo:
    st.image(LOGO_BYTES, width=180)
with col_title:
    st.markdown("## Fiyat Teklifi Oluşturucu")

st.divider()

with st.sidebar:
    st.markdown("### ⚙️ İletişim Ayarları")
    if "ayarlar" not in st.session_state:
        st.session_state.ayarlar = ayarlari_yukle()
    a = st.session_state.ayarlar
    a["iletisim_ad"]    = st.text_input("Ad Soyad",  a.get("iletisim_ad",""))
    a["iletisim_tel"]   = st.text_input("Telefon",   a.get("iletisim_tel",""))
    a["iletisim_email"] = st.text_input("E-posta",   a.get("iletisim_email",""))
    if st.button("💾 Kaydet"):
        ayarlari_kaydet(a)
        st.success("Kaydedildi!")

col_sol, col_sag = st.columns([1.2, 1])

with col_sol:
    st.markdown("#### 📋 Ürün Listesi")
    metin = st.text_area("Ürün listesini yapıştırın", height=400, label_visibility="collapsed")

with col_sag:
    st.markdown("#### ⚙️ Teklif Bilgileri")
    musteri_adi   = st.text_input("Müşteri Adı")
    musteri_sehir = st.text_input("Müşteri Şehri")
    teklif_no     = st.text_input("Teklif No")
    teklif_tarihi = st.text_input("Teklif Tarihi", value=date.today().strftime("%d/%m/%Y"))
    teslim_suresi = st.text_input("Teslim Süresi")
    teslimat      = st.text_input("Teslimat", value="Müşteri Tesisleri")
    indirim_str   = st.text_input("İndirim (€)", value="")
    st.markdown("---")
    not1 = st.text_input("Alt Not 1", value="K.D.V. hariçtir.")
    not2 = st.text_input("Alt Not 2", value="Fiyatlarımız EURO cinsindendir.")
    not3 = st.text_input("Alt Not 3", value="Ödeme: Peşin")
    not4 = st.text_input("Alt Not 4", value="")

def iletisim_metni():
    a = st.session_state.get("ayarlar", {})
    return "\n".join([v for v in [a.get("iletisim_ad",""),
                                    a.get("iletisim_tel",""),
                                    a.get("iletisim_email","")] if v])

if st.button("🔍 Ürünleri Önizle / Adet Düzenle", use_container_width=True):
    if not metin.strip():
        st.warning("Önce ürün listesini yapıştırın.")
    else:
        urunler = urunu_ayir(metin)
        if not urunler:
            st.error("Hiç ürün bulunamadı.")
        else:
            st.session_state["urunler"] = urunler
            st.success(f"{len(urunler)} ürün bulundu.")

if "urunler" in st.session_state:
    st.markdown("#### Ürün Listesi — Adet Düzenle")
    urunler = st.session_state["urunler"]
    for i, urun in enumerate(urunler):
        c1, c2, c3 = st.columns([3, 1, 1])
        with c1:
            st.text(urun["aciklama"].splitlines()[0][:60])
        with c2:
            st.text(f"{urun['fiyat']:,.2f} €")
        with c3:
            urun["adet"] = st.number_input("Adet", min_value=1,
                                            value=urun.get("adet",1),
                                            key=f"adet_{i}",
                                            label_visibility="collapsed")
    st.session_state["urunler"] = urunler

st.divider()

def get_kw():
    indirim = 0.0
    try:
        if indirim_str.strip():
            indirim = float(indirim_str.replace(",","."))
    except Exception:
        pass
    urunler = st.session_state.get("urunler",
                urunu_ayir(metin) if metin.strip() else [])
    return dict(
        urunler=urunler, musteri_adi=musteri_adi,
        musteri_sehir=musteri_sehir, teklif_no=teklif_no,
        teklif_tarihi=teklif_tarihi, teslim_suresi=teslim_suresi,
        teslimat=teslimat,
        alt_notlar=[n for n in [not1,not2,not3,not4] if n.strip()],
        iletisim=iletisim_metni(), indirim=indirim,
    )

musteri_dosya = re.sub(r'[\\/:*?"<>|]', "_", musteri_adi or "Teklif")

col_b1, col_b2, col_b3 = st.columns(3)

with col_b1:
    if st.button("📊 Excel Oluştur", use_container_width=True, type="primary"):
        kw = get_kw()
        if not kw["urunler"]:
            st.error("Hiç ürün bulunamadı.")
        else:
            buf = io.BytesIO()
            excel_olustur(**kw, cikti=buf)
            buf.seek(0)
            st.download_button("⬇️ Excel İndir", buf,
                file_name=f"Teklif_{musteri_dosya}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

with col_b2:
    if st.button("📄 PDF Oluştur", use_container_width=True, type="primary"):
        kw = get_kw()
        if not kw["urunler"]:
            st.error("Hiç ürün bulunamadı.")
        elif not HAS_PDF:
            st.error("reportlab kurulu değil.")
        else:
            buf = io.BytesIO()
            pdf_olustur(**kw, cikti=buf)
            buf.seek(0)
            st.download_button("⬇️ PDF İndir", buf,
                file_name=f"Teklif_{musteri_dosya}.pdf",
                mime="application/pdf",
                use_container_width=True)

with col_b3:
    if st.button("📦 Excel + PDF", use_container_width=True):
        kw = get_kw()
        if not kw["urunler"]:
            st.error("Hiç ürün bulunamadı.")
        else:
            buf_x = io.BytesIO()
            excel_olustur(**kw, cikti=buf_x)
            buf_x.seek(0)
            st.download_button("⬇️ Excel İndir", buf_x,
                file_name=f"Teklif_{musteri_dosya}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            if HAS_PDF:
                buf_p = io.BytesIO()
                pdf_olustur(**kw, cikti=buf_p)
                buf_p.seek(0)
                st.download_button("⬇️ PDF İndir", buf_p,
                    file_name=f"Teklif_{musteri_dosya}.pdf",
                    mime="application/pdf",
                    use_container_width=True)
